from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Q
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Data, synop
from user_account.models import Station
from datetime import datetime
from openpyxl import Workbook


@login_required(login_url="/")
def GetDashboard(request):
    context = {}
    today = datetime.today()
    synops = synop.objects.filter(created__year=today.year, created__month=today.month, created__day=today.day).select_related('data')[0:6]
    context['synops'] = synops
    return render(request, 'dashboard/index.html', context)


@login_required(login_url="/")
def GetStations(request):
    context = {}
    met_stations = Station.objects.prefetch_related('station').all()
    context['stations'] = met_stations
    return render(request, 'dashboard/stations.html', context)


@login_required(login_url="/")
def GetSynops(request):
    context = {}
    synops = synop.objects.all()
    context['synops'] = synops
    return render(request, 'dashboard/synops.html')

@login_required(login_url="/")
def GetDataExporting(request):
    context = {}
    stations = Station.objects.all()
    if request.method == 'POST':
        station = Station.objects.get(station_number=request.POST.get('station'))
        start = request.POST.get('start')
        end = request.POST.get('end')

        # format date string correctly
        start = start.replace('/', '-')
        end = end.replace('/', '-')
        
        #converts date string from above to datetime object
        formated_start_date = datetime.strptime(start, '%m-%d-%Y')
        formated_end_date = datetime.strptime(end, '%m-%d-%Y')

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] =  f'attachment; filename={station.station_name}.xlsx'

        data = Data.objects.filter(Q(station_id=station) and Q(created_at__range=[formated_start_date, formated_end_date])).values_list('created_at', request.POST.get('parameter'))
        
        if len(data) == 0:
            messages.error(request, "Data for the period specified not available")

        workbook = Workbook()
        worksheet = workbook.active

        columns = [
            request.POST.get('parameter'),
            'date',
        ]
        row_num = 1

        #assign columns title in excel sheet
        for col_num, col_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = col_title

        # assigning
        for item in data:
            row_num += 1

            rows = [
                item.dry_bulb if request.POST.get('parameter') == 'dry_bulb' else item.rainfall_amount,
                item.created_at
            ]

            for col_num, cell_value in enumerate(rows, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response
    context['stations'] = stations
    return render(request, "dashboard/data_exporting.html", context)